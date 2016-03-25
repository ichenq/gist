#!/usr/bin/python
# -*-coding:utf-8-*-

import os
import codecs
import openpyxl
import xml.etree.ElementTree as ET


cpp_template = """#pragma once

#include <stdint.h>

//Auto-generated code, DO NOT EDIT!
enum LanguageId : uint32_t
{
"""

go_template = """package proto

// Auto-generated code, DO NOT EDIT!
const (
"""

# 读取excel
def parse_excel(filepath):
    table = {}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet in wb:
        for n in range(2, sheet.max_row+1):
            name = sheet.cell(row=n, column=1).value
            code = int(sheet.cell(row=n, column=2).value)
            zh_CN = sheet.cell(row=n, column=3).value
            en_US = sheet.cell(row=n, column=4).value
            item = {
                "Enum": name,
                "zhCN": zh_CN,
                "enUS": en_US,
                #"zhTW"
            }
            assert code not in table
            table[code] = item
        return table


# xml换行
def indent(elem, level=0):
    i = "\n" + level*"    "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "    "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


# 生成xml
def write_xml(table, filepath):
    root = ET.Element('config')
    keys = table.keys()
    keys.sort()
    for ec in keys:
        item = table[ec]
        elem = ET.SubElement(root, 'Error')
        #elem.set('enum', item['Enum'])
        elem.set('code', str(ec))
        elem.set('lang_zhCN', item['zhCN'])
        elem.set('lang_enUS', item['enUS'])
    indent(root)
    tree = ET.ElementTree(root)
    tree.write(filepath, encoding='utf-8', xml_declaration=True)


# 计算最大长度
def calc_max_len(table):
    max_len = 0
    for k, item in table.items():
        name = item['Enum']
        if len(name) > max_len:
            max_len = len(name)
    return max_len


# 空格对齐
def space_padding(text, min_len):
    if len(text) < min_len:
        for n in range(min_len - len(text)):
            text += ' '
    return text


# 生成go源文件
def gen_go_src(table, max_len, filepath):
    content = go_template
    keys = table.keys()
    keys.sort()
    for ec in keys:
        item = table[ec]
        name = item['Enum']
        name = space_padding(name, max_len)
        value = space_padding(str(ec), 6)
        line = '\t%s uint32 = %s //%s\n' % (name, value, item['zhCN'])
        content += line
    content += "\n)"
    f = codecs.open(filepath, "w", "utf-8")
    f.write(content)
    f.close()


# 生成C++头文件
def gen_cpp_src(table, max_len, filepath):
    content = cpp_template
    keys = table.keys()
    keys.sort()
    for ec in keys:
        item = table[ec]
        name = item['Enum']
        name = space_padding(name, max_len)
        value = space_padding(str(ec), 6)
        line = "    %s = %s,  // %s\n" % (name, value, item['zhCN'])
        content += line
    content += "\n};"
    f = codecs.open(filepath, "w", "utf-8")
    f.write(content)
    f.close()


def main(excelpath, xmlpath, gofile, cppfile):
    table = parse_excel(excelpath)
    max_len = calc_max_len(table)
    write_xml(table, xmlpath)
    gen_go_src(table, max_len, gofile)
    gen_cpp_src(table, max_len, cppfile)


if __name__ == '__main__':
    excelpath = u"多语言错误码.xlsx"
    xmlpath = u"networkError.xml"
    gofile = u"errno.go"
    cppfile = u"LanguageId.h"
    main(excelpath, xmlpath, gofile, cppfile)
